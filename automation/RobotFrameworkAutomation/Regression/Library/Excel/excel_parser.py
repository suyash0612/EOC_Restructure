import openpyxl
import json
import pandas as pd

def Parse_Excel_File(file_path, sheet_name, selected_column):
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    
    data_dict = {}

    # Check if the selected column exists in the DataFrame
    if selected_column not in df.columns:
        print(f"Column '{selected_column}' not found in the sheet.")
        return data_dict

    # Iterate through the rows of the DataFrame
    for index, row in df.iterrows():
        field = row[0]  # Assuming the first column is 'field'
        dynamic_variable = row[1]  # Assuming the second column is 'dynamic_variable'
        column_value = row[selected_column]  # Access the selected column value

        if pd.notna(column_value):  # Check if the value is not NaN
            # Convert column_value to int if it's numeric, otherwise keep it as is
            column_value = int(column_value) if isinstance(column_value, (int, float)) else str(column_value)

            # Only add the entry if the field is not empty
            if field not in data_dict:
                data_dict[field] = []
            data_dict[field].append([dynamic_variable, column_value])

    return data_dict

def Update_Excel_Value(file_path, sheet_name, selected_column, field_to_match, new_value):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # Find the column index based on the selected_column
    column_index = None
    headers = sheet[1]

    for col_idx, header in enumerate(headers, start=1):
        if str(header.value).strip() == str(selected_column).strip():
            column_index = col_idx
            print(column_index)
            break

    if column_index is None:
        print(f"Column '{selected_column}' not found in the sheet.")
        return

    # Find the row index based on the field_to_match value
    row_index = None
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == field_to_match:  # Assuming Field is in the first column
            row_index = idx
            print(row_index)
            break

    if row_index is None:
        print(f"No row found with '{field_to_match}' in the Field column.")
        return

    # Update the cell value
    print(new_value)
    sheet.cell(row=row_index, column=column_index, value=json.dumps(new_value))

    # Save the workbook
    workbook.save(file_path)

def Update_SRId_Into_Excel_For_Testcases(file_path, sheet_name, selected_column,field_to_match, service,new_value):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # Find the column index based on the selected_column
    column_index = None
    headers = sheet[1]

    for col_idx, header in enumerate(headers, start=1):
        if str(header.value).strip() == str(selected_column).strip():
            column_index = col_idx
            # print(column_index)
            break

    if column_index is None:
        print(f"Column '{selected_column}' not found in the sheet.")
        return

    # Find the row index based on the field_to_match value
    row_index = None
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == field_to_match:  # Assuming Field is in the first column
            row_index = idx
            # print(row_index)
            break

    if row_index is None:
        print(f"No row found with '{field_to_match}' in the Field column.")
        return

    # Update the cell value
    print("Value Updated")
    sheet.cell(row=row_index, column=column_index, value=json.dumps(new_value))

    # Save the workbook
    workbook.save(file_path)

def Add_New_Field_And_Value_To_Excel(file_path, sheet_name, selected_column, new_field, new_value):
    # Load the Excel file into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')

    # Check if required columns exist
    if 'Field' not in df.columns or selected_column not in df.columns:
        print(f"Required column 'Field' or '{selected_column}' not found in the sheet.")
        return

    # Search if new_field already exists in the 'Field' column
    if new_field in df['Field'].values:
        print(f"Field '{new_field}' already exists. Updating its value.")
        row_to_update = df[df['Field'] == new_field].index[0]  # Get the index of the existing field

        # Update the selected column with the new value
        df.at[row_to_update, selected_column] = json.dumps(new_value) if isinstance(new_value, dict) else new_value

        # Update DynamicVariable.fieldName if the column exists
        if 'Dynamic Variable' in df.columns:
            df.at[row_to_update, 'Dynamic Variable'] = f"DynamicVariable.{new_field}"
    else:
        # If the field doesn't exist, add it as a new row
        print(f"Field '{new_field}' does not exist. Adding it as a new field.")
        
        # Create a new DataFrame for the new row
        new_row = pd.DataFrame({
            'Field': [new_field],
            selected_column: [json.dumps(new_value) if isinstance(new_value, dict) else new_value],
            'Dynamic Variable': [f"DynamicVariable.{new_field}"] if 'Dynamic Variable' in df.columns else [None]
        })

        # Concatenate the new row to the existing DataFrame
        df = pd.concat([df, new_row], ignore_index=True)  # Append the new row to the DataFrame

    # Save the DataFrame back to the Excel file
    df.to_excel(file_path, sheet_name=sheet_name, index=False, engine='openpyxl')
    print(f"Field '{new_field}' with value '{new_value}' processed successfully.")


